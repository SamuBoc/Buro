from datetime import datetime, timedelta
import io
import json
import logging
import mimetypes
import os
import urllib.request
import uuid

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.db.models import Case as DjangoCase, Count, IntegerField, Q, Value, When
from django.http import FileResponse, Http404, HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from accounts.constants import ROLE_ADMINISTRADOR, ROLE_ESTUDIANTE, ROLE_PROFESOR, ROLE_SECRETARIA
from accounts.decorators import role_required
from accounts.permissions import (
    can_access_recording,
    can_add_interaction,
    can_manage_case_deadline,
    can_reassign_case,
    can_view_case,
)
from core.utils import get_client_ip

from .email_utils import send_interaction_email
from .forms import (
    CaseDeadlineForm,
    CaseForm,
    CaseReassignmentForm,
    CaseRejectionForm,
    CommunicationInteractionForm,
)
from .models import Case, CaseAuditLog, CallSession, CaseDocument, CommunicationInteraction, Notification
from .services import auto_assign_case, reassign_case

User = get_user_model()


def _get_user_draft(user):
    return (
        Case.objects
        .filter(created_by=user, status=Case.STATUS_DRAFT)
        .order_by('-created_at', '-pk')
        .first()
    )


def _can_access_draft(user, case):
    return user.is_superuser or case.created_by_id == user.id


def _render_case_form(request, form, draft_case=None, is_editing_draft=False):
    return render(request, 'cases/case_register.html', {
        'form': form,
        'draft_case': draft_case,
        'is_editing_draft': is_editing_draft,
    })


def _enforce_complete_case_requirements(form):
    required_fields = {
        'sala': 'Este campo es obligatorio.',
        'description': 'Este campo es obligatorio.',
        'beneficiary': 'Este campo es obligatorio.',
    }
    has_errors = False
    for field_name, message in required_fields.items():
        value = form.cleaned_data.get(field_name)
        if value in (None, ''):
            form.add_error(field_name, message)
            has_errors = True

    documents = form.cleaned_data.get('documents') or []
    if not documents:
        form.add_error('documents', 'Debe cargar al menos un documento para el caso.')
        has_errors = True

    return not has_errors


def _build_deadline_priority(case, today):
    if not case.deadline_date:
        return {
            'text': 'Sin fecha limite',
            'days_remaining': None,
            'priority_label': 'Sin prioridad',
            'priority_class': 'priority-none',
            'deadline_class': 'deadline-none',
        }

    days_remaining = (case.deadline_date - today).days

    if days_remaining < 0:
        priority_label = 'Critica'
        priority_class = 'priority-critical'
        deadline_text = f'Vencido hace {abs(days_remaining)} dia(s)'
        deadline_class = 'deadline-overdue'
    elif days_remaining <= 2:
        priority_label = 'Alta'
        priority_class = 'priority-high'
        deadline_text = 'Vence hoy' if days_remaining == 0 else f'Vence en {days_remaining} dia(s)'
        deadline_class = 'deadline-soon'
    elif days_remaining <= 7:
        priority_label = 'Media'
        priority_class = 'priority-medium'
        deadline_text = f'Vence en {days_remaining} dia(s)'
        deadline_class = 'deadline-normal'
    else:
        priority_label = 'Baja'
        priority_class = 'priority-low'
        deadline_text = f'Vence en {days_remaining} dia(s)'
        deadline_class = 'deadline-normal'

    return {
        'text': deadline_text,
        'days_remaining': days_remaining,
        'priority_label': priority_label,
        'priority_class': priority_class,
        'deadline_class': deadline_class,
    }


def _build_academic_dashboard_filters(request):
    today = timezone.localdate()
    filtro_desde = request.GET.get('desde')
    filtro_hasta = request.GET.get('hasta')
    filtro_estado = request.GET.get('estado')
    filtro_sala = request.GET.get('sala')
    filtro_estudiante = (request.GET.get('estudiante') or '').strip()
    filtro_profesor = (request.GET.get('profesor') or '').strip()
    desde_date = None
    hasta_date = None
    student_filter = None
    professor_filter = None
    students_for_filter = (
        User.objects
        .filter(is_active=True, groups__name=ROLE_ESTUDIANTE)
        .order_by('first_name', 'last_name', 'username')
        .distinct()
    )
    professors_for_filter = (
        User.objects
        .filter(is_active=True, groups__name=ROLE_PROFESOR)
        .order_by('first_name', 'last_name', 'username')
        .distinct()
    )

    case_filters = Q(assigned_student__isnull=False)
    assigned_cases_filters = Q(assigned_cases__isnull=False)
    if filtro_estado:
        case_filters &= Q(state=filtro_estado)
        assigned_cases_filters &= Q(assigned_cases__state=filtro_estado)
    if filtro_sala:
        case_filters &= Q(sala=filtro_sala)
        assigned_cases_filters &= Q(assigned_cases__sala=filtro_sala)
    if filtro_desde:
        try:
            desde_date = datetime.strptime(filtro_desde, '%Y-%m-%d').date()
            case_filters &= Q(created_at__date__gte=desde_date)
            assigned_cases_filters &= Q(assigned_cases__created_at__date__gte=desde_date)
        except ValueError:
            filtro_desde = ''
            desde_date = None
    if filtro_hasta:
        try:
            hasta_date = datetime.strptime(filtro_hasta, '%Y-%m-%d').date()
            case_filters &= Q(created_at__date__lte=hasta_date)
            assigned_cases_filters &= Q(assigned_cases__created_at__date__lte=hasta_date)
        except ValueError:
            filtro_hasta = ''
            hasta_date = None
    if filtro_estudiante:
        try:
            student_filter = int(filtro_estudiante)
            case_filters &= Q(assigned_student_id=student_filter)
            assigned_cases_filters &= Q(assigned_cases__assigned_student_id=student_filter)
        except ValueError:
            filtro_estudiante = ''
            student_filter = None
    if filtro_profesor:
        try:
            professor_filter = int(filtro_profesor)
            case_filters &= Q(assigned_student__profile__supervising_professor_id=professor_filter)
            assigned_cases_filters &= Q(
                assigned_cases__assigned_student__profile__supervising_professor_id=professor_filter
            )
        except ValueError:
            filtro_profesor = ''
            professor_filter = None

    sala_label = ''
    if filtro_sala:
        for value, label in Case.ROOM_CHOICES:
            if value == filtro_sala:
                sala_label = label
                break

    date_range_label = ''
    if desde_date and hasta_date:
        date_range_label = f'Desde {desde_date:%d/%m/%Y} hasta {hasta_date:%d/%m/%Y}'
    elif desde_date:
        date_range_label = f'Desde {desde_date:%d/%m/%Y}'
    elif hasta_date:
        date_range_label = f'Hasta {hasta_date:%d/%m/%Y}'

    return {
        'today': today,
        'filtro_desde': filtro_desde,
        'filtro_hasta': filtro_hasta,
        'filtro_estado': filtro_estado,
        'filtro_sala': filtro_sala,
        'filtro_estudiante': filtro_estudiante,
        'filtro_profesor': filtro_profesor,
        'student_filter': student_filter,
        'professor_filter': professor_filter,
        'students_for_filter': students_for_filter,
        'professors_for_filter': professors_for_filter,
        'sala_label': sala_label,
        'date_range_label': date_range_label,
        'case_filters': case_filters,
        'assigned_cases_filters': assigned_cases_filters,
        'filter_query': request.GET.urlencode(),
    }


@login_required
def case_list(request):
    today = timezone.localdate()
    search_query = (request.GET.get('q') or '').strip()
    filtro_sala = request.GET.get('sala') or ''
    filtro_estado = request.GET.get('estado') or ''
    filtro_prioridad = request.GET.get('prioridad') or ''
    orden = request.GET.get('orden') or 'fecha_desc'
    filtro_desde = request.GET.get('desde') or ''
    filtro_hasta = request.GET.get('hasta') or ''

    cases_qs = Case.objects.select_related('beneficiary', 'assigned_student').all()

    if search_query:
        cases_qs = cases_qs.filter(
            Q(code__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(beneficiary__name__icontains=search_query)
            | Q(assigned_student__username__icontains=search_query)
            | Q(assigned_student__first_name__icontains=search_query)
            | Q(assigned_student__last_name__icontains=search_query)
        )

    if filtro_sala:
        cases_qs = cases_qs.filter(sala=filtro_sala)

    if filtro_estado:
        cases_qs = cases_qs.filter(state=filtro_estado)

    if filtro_prioridad:
        if filtro_prioridad == 'none':
            cases_qs = cases_qs.filter(deadline_date__isnull=True)
        elif filtro_prioridad == 'critical':
            cases_qs = cases_qs.filter(deadline_date__lt=today)
        elif filtro_prioridad == 'high':
            cases_qs = cases_qs.filter(
                deadline_date__gte=today,
                deadline_date__lte=today + timedelta(days=2),
            )
        elif filtro_prioridad == 'medium':
            cases_qs = cases_qs.filter(
                deadline_date__gt=today + timedelta(days=2),
                deadline_date__lte=today + timedelta(days=7),
            )
        elif filtro_prioridad == 'low':
            cases_qs = cases_qs.filter(deadline_date__gt=today + timedelta(days=7))

    if filtro_desde:
        try:
            desde_date = datetime.strptime(filtro_desde, '%Y-%m-%d').date()
            cases_qs = cases_qs.filter(created_at__date__gte=desde_date)
        except ValueError:
            filtro_desde = ''

    if filtro_hasta:
        try:
            hasta_date = datetime.strptime(filtro_hasta, '%Y-%m-%d').date()
            cases_qs = cases_qs.filter(created_at__date__lte=hasta_date)
        except ValueError:
            filtro_hasta = ''

    order_map = {
        'fecha_desc': '-created_at',
        'fecha_asc': 'created_at',
        'vencimiento_asc': 'deadline_date',
        'vencimiento_desc': '-deadline_date',
        'codigo_asc': 'code',
        'codigo_desc': '-code',
    }

    if orden in {'prioridad_desc', 'prioridad_asc'}:
        cases_qs = cases_qs.annotate(
            priority_rank=DjangoCase(
                When(deadline_date__lt=today, then=Value(1)),
                When(deadline_date__gte=today, deadline_date__lte=today + timedelta(days=2), then=Value(2)),
                When(deadline_date__gt=today + timedelta(days=2), deadline_date__lte=today + timedelta(days=7), then=Value(3)),
                When(deadline_date__gt=today + timedelta(days=7), then=Value(4)),
                default=Value(5),
                output_field=IntegerField(),
            )
        )
        if orden == 'prioridad_desc':
            cases_qs = cases_qs.order_by('priority_rank', '-created_at')
        else:
            cases_qs = cases_qs.order_by('-priority_rank', '-created_at')
    else:
        order_by = order_map.get(orden, '-created_at')
        cases_qs = cases_qs.order_by(order_by)

    total_cases = cases_qs.count()
    state_choices = (
        Case.objects
        .order_by('state')
        .values_list('state', flat=True)
        .distinct()
    )
    cases = list(cases_qs)

    for case in cases:
        deadline_priority = _build_deadline_priority(case, today)
        case.deadline_status_text = deadline_priority['text']
        case.days_remaining       = deadline_priority['days_remaining']
        case.priority_label       = deadline_priority['priority_label']
        case.priority_class       = deadline_priority['priority_class']
        case.deadline_class       = deadline_priority['deadline_class']

    return render(request, 'cases/case_list.html', {
        'cases': cases,
        'total_cases': total_cases,
        'search_query': search_query,
        'filtro_sala': filtro_sala,
        'filtro_estado': filtro_estado,
        'filtro_prioridad': filtro_prioridad,
        'orden': orden,
        'filtro_desde': filtro_desde,
        'filtro_hasta': filtro_hasta,
        'state_choices': state_choices,
        'room_choices': Case.ROOM_CHOICES,
        'priority_choices': [
            {'value': 'critical', 'label': 'Critica'},
            {'value': 'high', 'label': 'Alta'},
            {'value': 'medium', 'label': 'Media'},
            {'value': 'low', 'label': 'Baja'},
            {'value': 'none', 'label': 'Sin fecha limite'},
        ],
        'order_choices': [
            {'value': 'fecha_desc', 'label': 'Fecha (mas reciente)'},
            {'value': 'fecha_asc', 'label': 'Fecha (mas antigua)'},
            {'value': 'vencimiento_asc', 'label': 'Vencimiento (mas proximo)'},
            {'value': 'vencimiento_desc', 'label': 'Vencimiento (mas lejano)'},
            {'value': 'prioridad_desc', 'label': 'Prioridad (critica primero)'},
            {'value': 'prioridad_asc', 'label': 'Prioridad (baja primero)'},
            {'value': 'codigo_asc', 'label': 'Codigo (A-Z)'},
            {'value': 'codigo_desc', 'label': 'Codigo (Z-A)'},
        ],
    })


@role_required(ROLE_PROFESOR, ROLE_ADMINISTRADOR)
def academic_dashboard(request):
    filters = _build_academic_dashboard_filters(request)
    today = filters['today']
    case_filters = filters['case_filters']
    assigned_cases_filters = filters['assigned_cases_filters']

    students = (
        User.objects
        .filter(is_active=True, groups__name=ROLE_ESTUDIANTE)
        .annotate(
            total_cases=Count(
                'assigned_cases',
                filter=assigned_cases_filters,
                distinct=True,
            ),
            overdue_cases=Count(
                'assigned_cases',
                filter=assigned_cases_filters & Q(assigned_cases__deadline_date__lt=today),
                distinct=True,
            ),
            without_deadline_cases=Count(
                'assigned_cases',
                filter=assigned_cases_filters & Q(assigned_cases__deadline_date__isnull=True),
                distinct=True,
            ),
        )
        .order_by('first_name', 'last_name', 'username')
    )
    if filters['student_filter']:
        students = students.filter(pk=filters['student_filter'])
    if filters['professor_filter']:
        students = students.filter(profile__supervising_professor_id=filters['professor_filter'])
    if filters['filtro_sala']:
        students = students.filter(total_cases__gt=0)

    cases_with_student = Case.objects.filter(case_filters)
    total_assigned = cases_with_student.count()
    total_overdue = cases_with_student.filter(deadline_date__lt=today).count()
    total_without_deadline = cases_with_student.filter(deadline_date__isnull=True).count()

    sala_counts = {
        row['sala']: row['count']
        for row in cases_with_student.values('sala').annotate(count=Count('id'))
    }
    sala_distribution = []
    for value, label in Case.ROOM_CHOICES:
        cantidad = sala_counts.get(value, 0)
        porcentaje = round((cantidad / total_assigned * 100), 1) if total_assigned else 0.0
        sala_distribution.append({
            'sala': label,
            'cantidad': cantidad,
            'porcentaje': porcentaje,
        })

    return render(request, 'cases/academic_dashboard.html', {
        'students': students,
        'total_students': students.count(),
        'total_assigned_cases': total_assigned,
        'total_overdue_cases': total_overdue,
        'total_without_deadline_cases': total_without_deadline,
        'sala_distribution': sala_distribution,
        'filtro_desde': filters['filtro_desde'],
        'filtro_hasta': filters['filtro_hasta'],
        'filtro_estado': filters['filtro_estado'],
        'filtro_sala': filters['filtro_sala'],
        'filtro_estudiante': filters['filtro_estudiante'],
        'filtro_profesor': filters['filtro_profesor'],
        'students_for_filter': filters['students_for_filter'],
        'professors_for_filter': filters['professors_for_filter'],
        'estado_choices': [
            Case.STATE_PENDING,
            Case.STATE_ASSIGNED,
            Case.STATE_NO_STUDENTS,
            Case.STATE_REJECTED,
        ],
        'sala_choices': Case.ROOM_CHOICES,
        'filter_query': filters['filter_query'],
    })


@role_required(ROLE_PROFESOR, ROLE_ADMINISTRADOR)
def academic_student_detail(request, student_id):
    today = timezone.localdate()
    filtro_desde = request.GET.get('desde')
    filtro_hasta = request.GET.get('hasta')
    filtro_estado = request.GET.get('estado')
    filtro_sala = request.GET.get('sala')
    desde_date = None
    hasta_date = None
    student = get_object_or_404(
        User.objects.filter(is_active=True, groups__name=ROLE_ESTUDIANTE),
        pk=student_id,
    )

    case_filters = Q(assigned_student=student)
    if filtro_estado:
        case_filters &= Q(state=filtro_estado)
    if filtro_sala:
        case_filters &= Q(sala=filtro_sala)
    if filtro_desde:
        try:
            desde_date = datetime.strptime(filtro_desde, '%Y-%m-%d').date()
            case_filters &= Q(created_at__date__gte=desde_date)
        except ValueError:
            filtro_desde = ''
            desde_date = None
    if filtro_hasta:
        try:
            hasta_date = datetime.strptime(filtro_hasta, '%Y-%m-%d').date()
            case_filters &= Q(created_at__date__lte=hasta_date)
        except ValueError:
            filtro_hasta = ''
            hasta_date = None

    sala_label = ''
    if filtro_sala:
        for value, label in Case.ROOM_CHOICES:
            if value == filtro_sala:
                sala_label = label
                break

    date_range_label = ''
    if desde_date and hasta_date:
        date_range_label = f'Desde {desde_date:%d/%m/%Y} hasta {hasta_date:%d/%m/%Y}'
    elif desde_date:
        date_range_label = f'Desde {desde_date:%d/%m/%Y}'
    elif hasta_date:
        date_range_label = f'Hasta {hasta_date:%d/%m/%Y}'

    student_cases = (
        Case.objects
        .select_related('beneficiary')
        .filter(case_filters)
        .order_by('-created_at')
    )

    metrics = {
        'total_cases': student_cases.count(),
        'overdue_cases': student_cases.filter(deadline_date__lt=today).count(),
        'without_deadline_cases': student_cases.filter(deadline_date__isnull=True).count(),
        'rejected_cases': student_cases.filter(state=Case.STATE_REJECTED).count(),
    }

    return render(request, 'cases/academic_student_detail.html', {
        'student': student,
        'student_cases': student_cases,
        'metrics': metrics,
        'filtro_desde': filtro_desde,
        'filtro_hasta': filtro_hasta,
        'filtro_estado': filtro_estado,
        'filtro_sala': filtro_sala,
        'sala_label': sala_label,
        'date_range_label': date_range_label,
        'estado_choices': [
            Case.STATE_PENDING,
            Case.STATE_ASSIGNED,
            Case.STATE_NO_STUDENTS,
            Case.STATE_REJECTED,
        ],
        'sala_choices': Case.ROOM_CHOICES,
        'filter_query': request.GET.urlencode(),
    })


@role_required(ROLE_SECRETARIA, ROLE_ADMINISTRADOR, ROLE_ESTUDIANTE)
def case_create(request):
    draft_case = _get_user_draft(request.user)

    if request.method == 'POST':
        submit_action = request.POST.get('submit_action', 'complete')
        is_draft_submission = submit_action == 'draft'
        form = CaseForm(
            request.POST,
            request.FILES,
            instance=draft_case,
            allow_partial=is_draft_submission,
        )

        if form.is_valid():
            try:
                with transaction.atomic():
                    case = form.save(commit=False)
                    case.created_by = request.user
                    case._request   = request
                    case.status = (
                        Case.STATUS_DRAFT if is_draft_submission else Case.STATUS_COMPLETE
                    )
                    case.save()

                    for document in form.cleaned_data.get('documents', []):
                        CaseDocument.objects.create(case=case, file=document)

                    student = None
                    if case.status == Case.STATUS_COMPLETE:
                        student = auto_assign_case(case)

                if is_draft_submission:
                    messages.success(
                        request,
                        f'Se guardo el borrador del caso {case.code}. Puedes continuarlo despues.'
                    )
                elif student is None:
                    messages.warning(
                        request,
                        f'El caso {case.code} fue registrado pero no hay estudiantes disponibles para asignarlo.'
                    )
                else:
                    messages.success(
                        request,
                        f'El caso {case.code} fue registrado y asignado a {student.get_full_name() or student.username}.'
                    )
                if is_draft_submission:
                    return redirect('case_create')
                return redirect('case_detail', pk=case.pk)
            except Exception:
                messages.error(
                    request,
                    'Ocurrio un problema al guardar el caso y sus documentos. Intente nuevamente.'
                )
        else:
            messages.error(request, 'Por favor corrija los errores del formulario.')
    else:
        form = CaseForm(instance=draft_case, allow_partial=bool(draft_case))
        if draft_case:
            messages.info(
                request,
                f'Tienes un borrador pendiente ({draft_case.code}). Puedes continuar completandolo.'
            )

    return _render_case_form(request, form, draft_case=draft_case)


@login_required
def case_draft_list(request):
    drafts = (
        Case.objects
        .filter(created_by=request.user, status=Case.STATUS_DRAFT)
        .select_related('beneficiary', 'assigned_student')
        .order_by('-created_at')
    )
    return render(request, 'cases/case_draft_list.html', {'drafts': drafts})


@login_required
def case_edit_draft(request, pk):
    draft_case = get_object_or_404(
        Case.objects.select_related('beneficiary', 'assigned_student'),
        pk=pk,
        status=Case.STATUS_DRAFT,
    )

    if not _can_access_draft(request.user, draft_case):
        messages.error(request, 'No tienes permisos para editar este borrador.')
        return redirect('case_draft_list')

    if request.method == 'POST':
        submit_action       = request.POST.get('submit_action', 'draft')
        is_draft_submission = submit_action == 'draft'
        form = CaseForm(
            request.POST,
            request.FILES,
            instance=draft_case,
            allow_partial=is_draft_submission,
        )

        if form.is_valid():
            try:
                with transaction.atomic():
                    case = form.save(commit=False)
                    case.created_by = draft_case.created_by or request.user
                    case._request   = request
                    case.status = (
                        Case.STATUS_DRAFT if is_draft_submission else Case.STATUS_COMPLETE
                    )
                    case.save()

                    for document in form.cleaned_data.get('documents', []):
                        CaseDocument.objects.create(case=case, file=document)

                    student = None
                    if case.status == Case.STATUS_COMPLETE:
                        student = auto_assign_case(case)

                if is_draft_submission:
                    messages.success(request, f'Se actualizo el borrador del caso {case.code}.')
                    return redirect('case_edit_draft', pk=case.pk)

                if student is None:
                    messages.warning(
                        request,
                        f'El caso {case.code} fue completado pero no hay estudiantes disponibles para asignarlo.'
                    )
                else:
                    messages.success(
                        request,
                        f'El caso {case.code} fue completado y asignado a {student.get_full_name() or student.username}.'
                    )
                return redirect('case_detail', pk=case.pk)
            except Exception:
                messages.error(
                    request,
                    'Ocurrio un problema al actualizar el borrador. Intente nuevamente.'
                )
        else:
            messages.error(request, 'Por favor corrija los errores del formulario.')
    else:
        form = CaseForm(instance=draft_case, allow_partial=True)

    return _render_case_form(request, form, draft_case=draft_case, is_editing_draft=True)


@login_required
def case_detail(request, pk):
    case = get_object_or_404(
        Case.objects
        .select_related('beneficiary', 'assigned_student')
        .prefetch_related(
            'documents',
            'reassignment_logs__changed_by',
            'reassignment_logs__old_student',
            'reassignment_logs__new_student',
            'interactions__registered_by',
        ),
        pk=pk,
    )

    if not can_view_case(request.user, case):
        messages.error(request, 'No tienes permisos para acceder a este caso.')
        return redirect('case_list')

    return render(request, 'cases/case_detail.html', {
        'case':                case,
        'can_reassign':        can_reassign_case(request.user),
        'can_manage_deadline': can_manage_case_deadline(request.user),
        'can_add_interaction':   can_add_interaction(request.user, case),
        'can_access_recording':  can_access_recording(request.user, case),
        'deadline_form':         CaseDeadlineForm(instance=case),
        'reassignment_form':   CaseReassignmentForm(case=case),
        'rejection_form':      CaseRejectionForm(instance=case),
        'interaction_form':    CommunicationInteractionForm(),
        'interactions':        case.interactions.select_related('registered_by').order_by('-timestamp'),
    })


@role_required(ROLE_SECRETARIA, ROLE_PROFESOR, ROLE_ADMINISTRADOR)
def case_update_deadline(request, pk):
    case = get_object_or_404(
        Case.objects.select_related('beneficiary', 'assigned_student'),
        pk=pk,
    )

    if request.method != 'POST':
        return redirect('case_detail', pk=case.pk)

    form = CaseDeadlineForm(request.POST, instance=case)

    if form.is_valid():
        deadline_case = form.save()
        messages.success(request, f'La fecha limite del caso {deadline_case.code} fue actualizada.')
        return redirect('case_detail', pk=case.pk)

    return render(request, 'cases/case_detail.html', {
        'case':                case,
        'can_reassign':        can_reassign_case(request.user),
        'can_manage_deadline': can_manage_case_deadline(request.user),
        'can_add_interaction': can_add_interaction(request.user, case),
        'deadline_form':       form,
        'reassignment_form':   CaseReassignmentForm(case=case),
        'rejection_form':      CaseRejectionForm(instance=case),
        'interaction_form':    CommunicationInteractionForm(),
        'interactions':        case.interactions.select_related('registered_by').order_by('-timestamp'),
    })


@role_required(ROLE_SECRETARIA, ROLE_PROFESOR, ROLE_ADMINISTRADOR)
def case_reassign(request, pk):
    case = get_object_or_404(
        Case.objects.select_related('beneficiary', 'assigned_student'),
        pk=pk,
    )

    if request.method != 'POST':
        return redirect('case_detail', pk=case.pk)

    form = CaseReassignmentForm(request.POST, case=case)

    if form.is_valid():
        new_student   = form.cleaned_data['assigned_student']
        old_student   = reassign_case(case, new_student, request.user)
        previous_name = old_student.get_full_name() or old_student.username if old_student else 'Sin asignar'
        new_name      = new_student.get_full_name() or new_student.username
        messages.success(
            request,
            f'El caso {case.code} fue reasignado de {previous_name} a {new_name}.'
        )
        return redirect('case_detail', pk=case.pk)

    messages.error(request, 'Por favor seleccione un estudiante valido.')
    return render(request, 'cases/case_detail.html', {
        'case':                case,
        'can_reassign':        can_reassign_case(request.user),
        'can_manage_deadline': can_manage_case_deadline(request.user),
        'can_add_interaction':   can_add_interaction(request.user, case),
        'can_access_recording':  can_access_recording(request.user, case),
        'deadline_form':         CaseDeadlineForm(instance=case),
        'reassignment_form':   form,
        'rejection_form':      CaseRejectionForm(instance=case),
        'interaction_form':    CommunicationInteractionForm(),
        'interactions':        case.interactions.select_related('registered_by').order_by('-timestamp'),
    })


@role_required(ROLE_SECRETARIA, ROLE_PROFESOR, ROLE_ADMINISTRADOR)
def case_reject(request, pk):
    case = get_object_or_404(
        Case.objects.select_related('beneficiary', 'assigned_student'),
        pk=pk,
    )

    if request.method != 'POST':
        return redirect('case_detail', pk=case.pk)

    if case.state == Case.STATE_REJECTED:
        messages.warning(request, 'Este caso ya fue rechazado anteriormente.')
        return redirect('case_detail', pk=case.pk)

    form = CaseRejectionForm(request.POST, instance=case)

    if form.is_valid():
        try:
            with transaction.atomic():
                case.state            = Case.STATE_REJECTED
                case.rejection_reason = form.cleaned_data['rejection_reason']
                case._request         = request
                case.save()
                messages.success(request, f'El caso {case.code} ha sido rechazado exitosamente.')
        except Exception:
            messages.error(request, 'Ocurrió un problema al rechazar el caso. Intente nuevamente.')
        return redirect('case_detail', pk=case.pk)

    messages.error(request, 'Por favor ingrese una causal de rechazo válida.')
    return render(request, 'cases/case_detail.html', {
        'case':                case,
        'can_reassign':        can_reassign_case(request.user),
        'can_manage_deadline': can_manage_case_deadline(request.user),
        'can_add_interaction':   can_add_interaction(request.user, case),
        'can_access_recording':  can_access_recording(request.user, case),
        'deadline_form':         CaseDeadlineForm(instance=case),
        'reassignment_form':   CaseReassignmentForm(case=case),
        'rejection_form':      form,
        'interaction_form':    CommunicationInteractionForm(),
        'interactions':        case.interactions.select_related('registered_by').order_by('-timestamp'),
    })


@login_required
def case_add_interaction(request, case_id):
    case = get_object_or_404(
        Case.objects.select_related('beneficiary', 'assigned_student'),
        pk=case_id,
    )

    if not can_add_interaction(request.user, case):
        messages.error(request, 'No tienes permiso para registrar interacciones en este caso.')
        return redirect('case_detail', pk=case_id)

    if request.method != 'POST':
        return redirect('case_detail', pk=case_id)

    form = CommunicationInteractionForm(request.POST)
    if form.is_valid():
        interaction               = form.save(commit=False)
        interaction.case          = case
        interaction.registered_by = request.user
        interaction.save()

        snippet = interaction.description[:100]
        if len(interaction.description) > 100:
            snippet += '...'

        CaseAuditLog.objects.create(
            case=case,
            user=request.user,
            action='COMMUNICATION',
            description=(
                f'{interaction.get_interaction_type_display()} '
                f'({interaction.get_direction_display()}) registrada en el caso '
                f'{case.code}: {snippet}'
            ),
            case_radicado=case.code,
            ip_address=get_client_ip(request),
        )

        if interaction.interaction_type == CommunicationInteraction.TYPE_EMAIL:
            send_interaction_email(interaction)
            messages.success(
                request,
                f'Interacción registrada y correo enviado a {case.beneficiary.email}.'
            )
        elif interaction.interaction_type == CommunicationInteraction.TYPE_MESSAGE:
            phone = getattr(case.beneficiary, 'phone', '') or ''
            clean_phone = phone.replace(' ', '').replace('-', '')
            whatsapp_url = f'https://wa.me/57{clean_phone}' if clean_phone else None
            if whatsapp_url:
                messages.success(
                    request,
                    f'Interacción registrada. Enlace WhatsApp: {whatsapp_url}'
                )
            else:
                messages.success(request, 'Interacción de comunicación registrada exitosamente.')
        else:
            messages.success(request, 'Interacción de comunicación registrada exitosamente.')
    else:
        messages.error(request, 'Por favor corrija los errores del formulario.')

    return redirect('case_detail', pk=case_id)


# ─── Audit logs ──────────────────────────────────────────────────────────────

@login_required
def case_audit_log(request, case_id):
    case = get_object_or_404(Case, pk=case_id)

    has_access = (
        request.user.is_staff
        or request.user.groups.filter(
            name__in=[ROLE_SECRETARIA, ROLE_PROFESOR, ROLE_ADMINISTRADOR]
        ).exists()
        or case.assigned_student == request.user
    )
    if not has_access:
        messages.error(request, 'No tienes permiso para ver la bitacora de este caso.')
        return redirect('case_list')

    logs = CaseAuditLog.objects.filter(case=case).select_related('user').order_by('-timestamp')

    return render(request, 'cases/case_audit_log.html', {
        'case':       case,
        'logs':       logs,
        'page_title': f'Bitacora - {case.code}',
    })


@login_required
def global_audit_log(request):
    if not (
        request.user.is_staff
        or request.user.groups.filter(name=ROLE_ADMINISTRADOR).exists()
    ):
        messages.error(request, 'Acceso restringido a administradores.')
        return redirect('case_list')

    logs = CaseAuditLog.objects.select_related('user', 'case').order_by('-timestamp')[:500]
    return render(request, 'cases/global_audit_log.html', {
        'logs':       logs,
        'page_title': 'Bitacora Global de Casos',
    })


@login_required
def notification_list(request):
    notifications = Notification.objects.filter(
        recipient_user=request.user
    ).select_related('case').order_by('-created_at')

    unread_count = notifications.filter(is_read=False).count()

    return render(request, 'cases/notifications.html', {
        'notifications': notifications,
        'unread_count':  unread_count,
        'page_title':    'Mis Notificaciones',
    })


@login_required
def mark_notification_read(request, notification_id):
    notification = get_object_or_404(
        Notification,
        pk=notification_id,
        recipient_user=request.user,
    )
    notification.mark_as_read()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'ok', 'notification_id': notification_id})

    return redirect('case_detail', pk=notification.case_id)


@login_required
def mark_all_notifications_read(request):
    if request.method == 'POST':
        Notification.objects.filter(
            recipient_user=request.user,
            is_read=False,
        ).update(is_read=True, read_at=timezone.now())
        messages.success(request, 'Todas las notificaciones marcadas como leidas.')
    return redirect('notification_list')


@login_required
def unread_notifications_count(request):
    count = Notification.objects.filter(
        recipient_user=request.user,
        is_read=False,
    ).count()
    return JsonResponse({'unread_count': count})


REPORT_KNOWN_STATES = [
    Case.STATE_PENDING,
    Case.STATE_ASSIGNED,
    Case.STATE_NO_STUDENTS,
    Case.STATE_REJECTED,
]


def _parse_report_date(raw_value):
    if not raw_value:
        return None
    try:
        return datetime.strptime(raw_value, '%Y-%m-%d').date()
    except ValueError:
        return None


@role_required(ROLE_ADMINISTRADOR)
def case_report_by_state(request):
    desde_raw = (request.GET.get('desde') or '').strip()
    hasta_raw = (request.GET.get('hasta') or '').strip()
    sala_raw  = (request.GET.get('sala') or '').strip()

    desde_date  = _parse_report_date(desde_raw)
    hasta_date  = _parse_report_date(hasta_raw)
    valid_salas = {value for value, _ in Case.ROOM_CHOICES}
    sala_filter = sala_raw if sala_raw in valid_salas else ''

    cases = Case.objects.all()
    if desde_date:
        cases = cases.filter(created_at__date__gte=desde_date)
    if hasta_date:
        cases = cases.filter(created_at__date__lte=hasta_date)
    if sala_filter:
        cases = cases.filter(sala=sala_filter)

    total        = cases.count()
    state_counts = {
        row['state']: row['count']
        for row in cases.values('state').annotate(count=Count('id'))
    }

    rows = []
    for state in REPORT_KNOWN_STATES:
        cantidad   = state_counts.pop(state, 0)
        porcentaje = round((cantidad / total * 100), 1) if total else 0.0
        rows.append({'estado': state, 'cantidad': cantidad, 'porcentaje': porcentaje})
    for extra_state, cantidad in state_counts.items():
        porcentaje = round((cantidad / total * 100), 1) if total else 0.0
        rows.append({'estado': extra_state or 'Sin estado', 'cantidad': cantidad, 'porcentaje': porcentaje})

    return render(request, 'cases/report_by_state.html', {
        'page_title':   'Reporte de casos por estado',
        'rows':         rows,
        'total':        total,
        'filtro_desde': desde_raw,
        'filtro_hasta': hasta_raw,
        'filtro_sala':  sala_filter,
        'salas':        Case.ROOM_CHOICES,
        'chart_labels': [row['estado']   for row in rows],
        'chart_values': [row['cantidad'] for row in rows],
    })


@role_required(ROLE_ADMINISTRADOR, ROLE_PROFESOR)
def case_report_by_sala(request):
    desde_raw  = (request.GET.get('desde') or '').strip()
    hasta_raw  = (request.GET.get('hasta') or '').strip()
    state_raw  = (request.GET.get('estado') or '').strip()

    desde_date   = _parse_report_date(desde_raw)
    hasta_date   = _parse_report_date(hasta_raw)
    valid_states = set(REPORT_KNOWN_STATES)
    state_filter = state_raw if state_raw in valid_states else ''

    cases = Case.objects.all()
    if desde_date:
        cases = cases.filter(created_at__date__gte=desde_date)
    if hasta_date:
        cases = cases.filter(created_at__date__lte=hasta_date)
    if state_filter:
        cases = cases.filter(state=state_filter)

    total      = cases.count()
    sala_counts = {
        row['sala']: row['count']
        for row in cases.values('sala').annotate(count=Count('id'))
    }

    rows = []
    for value, label in Case.ROOM_CHOICES:
        cantidad   = sala_counts.get(value, 0)
        porcentaje = round((cantidad / total * 100), 1) if total else 0.0
        rows.append({'sala': label, 'cantidad': cantidad, 'porcentaje': porcentaje})

    return render(request, 'cases/report_by_sala.html', {
        'page_title':    'Reporte de casos por sala jurídica',
        'rows':          rows,
        'total':         total,
        'filtro_desde':  desde_raw,
        'filtro_hasta':  hasta_raw,
        'filtro_estado': state_filter,
        'estados':       REPORT_KNOWN_STATES,
        'chart_labels':  [row['sala']     for row in rows],
        'chart_values':  [row['cantidad'] for row in rows],
    })


@role_required(ROLE_ADMINISTRADOR, ROLE_PROFESOR)
def export_cases_excel(request):
    desde_raw = (request.GET.get('desde') or '').strip()
    hasta_raw = (request.GET.get('hasta') or '').strip()
    sala_raw  = (request.GET.get('sala')  or '').strip()

    valid_salas = {value for value, _ in Case.ROOM_CHOICES}

    cases = Case.objects.select_related(
        'beneficiary', 'assigned_student'
    ).filter(status=Case.STATUS_COMPLETE)

    if _parse_report_date(desde_raw):
        cases = cases.filter(created_at__date__gte=_parse_report_date(desde_raw))
    if _parse_report_date(hasta_raw):
        cases = cases.filter(created_at__date__lte=_parse_report_date(hasta_raw))
    if sala_raw in valid_salas:
        cases = cases.filter(sala=sala_raw)

    cases = cases.order_by('-created_at')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Casos'

    headers = [
        'Código', 'Sala', 'Beneficiario', 'Estudiante Asignado',
        'Estado', 'Fecha de Creación', 'Fecha Límite',
    ]
    header_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, case in enumerate(cases, start=2):
        ws.cell(row=row_idx, column=1).value = case.code
        ws.cell(row=row_idx, column=2).value = case.get_sala_display() if case.sala else 'Sin sala'
        ws.cell(row=row_idx, column=3).value = case.beneficiary.name if case.beneficiary else 'Sin beneficiario'
        ws.cell(row=row_idx, column=4).value = (
            case.assigned_student.get_full_name() or case.assigned_student.username
            if case.assigned_student else 'Sin asignar'
        )
        ws.cell(row=row_idx, column=5).value = case.state
        ws.cell(row=row_idx, column=6).value = case.created_at.strftime('%d/%m/%Y')
        ws.cell(row=row_idx, column=7).value = (
            case.deadline_date.strftime('%d/%m/%Y') if case.deadline_date else '—'
        )

    column_widths = [15, 12, 25, 25, 35, 18, 15]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_casos.xlsx"'
    return response


@role_required(ROLE_ADMINISTRADOR, ROLE_PROFESOR)
def export_cases_pdf(request):
    desde_raw = (request.GET.get('desde') or '').strip()
    hasta_raw = (request.GET.get('hasta') or '').strip()
    sala_raw  = (request.GET.get('sala')  or '').strip()

    valid_salas = {value for value, _ in Case.ROOM_CHOICES}

    cases = Case.objects.select_related(
        'beneficiary', 'assigned_student'
    ).filter(status=Case.STATUS_COMPLETE)

    if _parse_report_date(desde_raw):
        cases = cases.filter(created_at__date__gte=_parse_report_date(desde_raw))
    if _parse_report_date(hasta_raw):
        cases = cases.filter(created_at__date__lte=_parse_report_date(hasta_raw))
    if sala_raw in valid_salas:
        cases = cases.filter(sala=sala_raw)

    cases = cases.order_by('-created_at')

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5 * inch, leftMargin=0.5 * inch,
        topMargin=0.5 * inch,   bottomMargin=0.5 * inch,
    )

    styles   = getSampleStyleSheet()
    elements = [
        Paragraph('<b>Reporte de Casos — Consultorio Jurídico ICESI</b>', styles['Title']),
        Spacer(1, 0.2 * inch),
    ]

    data = [['Código', 'Sala', 'Beneficiario', 'Estudiante Asignado', 'Estado', 'Fecha Creación']]
    for case in cases:
        data.append([
            case.code,
            case.get_sala_display() if case.sala else 'Sin sala',
            case.beneficiary.name if case.beneficiary else 'Sin beneficiario',
            (
                case.assigned_student.get_full_name() or case.assigned_student.username
                if case.assigned_student else 'Sin asignar'
            ),
            case.state,
            case.created_at.strftime('%d/%m/%Y'),
        ])

    table = Table(data, colWidths=[1.2*inch, 1*inch, 2*inch, 2*inch, 2.5*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#1A3A5C')),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0, 0), (-1, 0),  9),
        ('ALIGN',          (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',       (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F6F9')]),
        ('GRID',           (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING',     (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_casos.pdf"'
    return response


@role_required(ROLE_ADMINISTRADOR, ROLE_PROFESOR)
def export_academic_dashboard_excel(request):
    filters = _build_academic_dashboard_filters(request)
    assigned_cases_filters = filters['assigned_cases_filters']

    students = (
        User.objects
        .filter(is_active=True, groups__name=ROLE_ESTUDIANTE)
        .annotate(
            total_cases=Count('assigned_cases', filter=assigned_cases_filters, distinct=True),
            overdue_cases=Count(
                'assigned_cases',
                filter=assigned_cases_filters & Q(assigned_cases__deadline_date__lt=filters['today']),
                distinct=True,
            ),
            without_deadline_cases=Count(
                'assigned_cases',
                filter=assigned_cases_filters & Q(assigned_cases__deadline_date__isnull=True),
                distinct=True,
            ),
        )
        .order_by('first_name', 'last_name', 'username')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Panel Academico'

    headers = ['Estudiante', 'Usuario', 'Casos asignados', 'Casos vencidos', 'Sin fecha limite']
    header_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, student in enumerate(students, start=2):
        ws.cell(row=row_idx, column=1).value = student.get_full_name() or student.username
        ws.cell(row=row_idx, column=2).value = student.username
        ws.cell(row=row_idx, column=3).value = student.total_cases
        ws.cell(row=row_idx, column=4).value = student.overdue_cases
        ws.cell(row=row_idx, column=5).value = student.without_deadline_cases

    for col, width in enumerate([28, 18, 18, 16, 18], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_panel_academico.xlsx"'
    return response


@role_required(ROLE_ADMINISTRADOR, ROLE_PROFESOR)
def export_academic_dashboard_pdf(request):
    filters = _build_academic_dashboard_filters(request)
    assigned_cases_filters = filters['assigned_cases_filters']

    students = (
        User.objects
        .filter(is_active=True, groups__name=ROLE_ESTUDIANTE)
        .annotate(
            total_cases=Count('assigned_cases', filter=assigned_cases_filters, distinct=True),
            overdue_cases=Count(
                'assigned_cases',
                filter=assigned_cases_filters & Q(assigned_cases__deadline_date__lt=filters['today']),
                distinct=True,
            ),
            without_deadline_cases=Count(
                'assigned_cases',
                filter=assigned_cases_filters & Q(assigned_cases__deadline_date__isnull=True),
                distinct=True,
            ),
        )
        .order_by('first_name', 'last_name', 'username')
    )

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5 * inch, leftMargin=0.5 * inch,
        topMargin=0.5 * inch,   bottomMargin=0.5 * inch,
    )

    styles   = getSampleStyleSheet()
    elements = [
        Paragraph('<b>Panel Academico — Consultorio Juridico ICESI</b>', styles['Title']),
        Spacer(1, 0.2 * inch),
    ]

    data = [['Estudiante', 'Usuario', 'Casos asignados', 'Casos vencidos', 'Sin fecha limite']]
    for student in students:
        data.append([
            student.get_full_name() or student.username,
            student.username,
            str(student.total_cases),
            str(student.overdue_cases),
            str(student.without_deadline_cases),
        ])

    table = Table(data, colWidths=[2.2*inch, 1.4*inch, 1.2*inch, 1.2*inch, 1.4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#1A3A5C')),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0, 0), (-1, 0),  9),
        ('ALIGN',          (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',       (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F6F9')]),
        ('GRID',           (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING',     (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_panel_academico.pdf"'
    return response


@login_required
def serve_case_document(request, document_id):
    """Sirve de forma segura archivos adjuntos a un caso, validando permisos."""
    document = get_object_or_404(CaseDocument, pk=document_id)
    case     = document.case

    if not can_view_case(request.user, case):
        CaseAuditLog.objects.create(
            case=case,
            user=request.user,
            action='SECURITY_DENIED',
            description=(
                f'Acceso denegado al archivo "{document.file.name}" '
                f'del caso {case.code} por el usuario {request.user.username}.'
            ),
            case_radicado=case.code,
            ip_address=get_client_ip(request),
        )
        messages.error(request, 'No tienes permiso para acceder a este archivo.')
        return redirect('case_list')

    file_path = document.file.path
    if not os.path.exists(file_path):
        raise Http404('El archivo no existe.')

    mime_type, _ = mimetypes.guess_type(file_path)
    response = FileResponse(
        open(file_path, 'rb'),
        content_type=mime_type or 'application/octet-stream',
    )
    response['Content-Disposition'] = (
        f'inline; filename="{os.path.basename(file_path)}"'
    )
    return response


# ─── HU-22: Grabar llamadas ───────────────────────────────────────────────────

logger = logging.getLogger(__name__)



@login_required
def upload_call_recording(request, case_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    case = get_object_or_404(Case, pk=case_id)

    if not can_add_interaction(request.user, case):
        return JsonResponse({'error': 'Sin permiso'}, status=403)

    audio = request.FILES.get('audio')
    if not audio:
        return JsonResponse({'error': 'No se recibió archivo de audio'}, status=400)

    description = request.POST.get('description', 'Llamada grabada desde la plataforma')

    try:
        interaction = CommunicationInteraction.objects.create(
            case=case,
            registered_by=request.user,
            interaction_type=CommunicationInteraction.TYPE_CALL,
            direction=CommunicationInteraction.DIRECTION_OUT,
            description=description,
            audio_file=audio,
        )
        CaseAuditLog.objects.create(
            case=case,
            user=request.user,
            action='COMMUNICATION',
            description=f'Grabación de llamada subida para el caso {case.code}.',
            case_radicado=case.code,
            ip_address=get_client_ip(request),
        )
        return JsonResponse({'status': 'ok', 'interaction_id': interaction.pk})
    except Exception as exc:
        logger.error('Error al guardar grabación de llamada para caso %s: %s', case.code, exc)
        CaseAuditLog.objects.create(
            case=case,
            user=request.user,
            action='COMMUNICATION',
            description=f'Error al guardar grabación de llamada para el caso {case.code}: {exc}',
            case_radicado=case.code,
            ip_address=get_client_ip(request),
        )
        return JsonResponse({'error': 'Error al guardar la grabación'}, status=500)
# ─── HU-24: Métricas de canales de comunicación ──────────────────────────────

@login_required
@role_required(ROLE_ADMINISTRADOR)
def communication_metrics(request):
    tipo_filter = request.GET.get('tipo', '')

    qs = CommunicationInteraction.objects.all()
    if tipo_filter:
        qs = qs.filter(interaction_type=tipo_filter)

    by_type = (
        CommunicationInteraction.objects
        .values('interaction_type')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    type_labels = dict(CommunicationInteraction.TYPE_CHOICES)
    metrics = [
        {
            'type': row['interaction_type'],
            'label': type_labels.get(row['interaction_type'], row['interaction_type']),
            'count': row['count'],
        }
        for row in by_type
    ]

    interactions = qs.select_related('case', 'registered_by').order_by('-timestamp')[:100]

    return render(request, 'cases/communication_metrics.html', {
        'metrics':      metrics,
        'interactions': interactions,
        'tipo_filter':  tipo_filter,
        'tipo_choices': CommunicationInteraction.TYPE_CHOICES,
        'total':        CommunicationInteraction.objects.count(),
    })
# ─── HU-23: Acceso controlado a grabaciones ──────────────────────────────────

@login_required
def serve_call_recording(request, interaction_id):
    interaction = get_object_or_404(CommunicationInteraction, pk=interaction_id)
    case = interaction.case

    if not can_access_recording(request.user, case):
        CaseAuditLog.objects.create(
            case=case,
            user=request.user,
            action='SECURITY_DENIED',
            description=(
                f'Acceso denegado a grabación de llamada del caso {case.code} '
                f'por el usuario {request.user.username}.'
            ),
            case_radicado=case.code,
            ip_address=get_client_ip(request),
        )
        return HttpResponseForbidden('No tienes permiso para acceder a esta grabación.')

    if not interaction.audio_file:
        raise Http404('No hay grabación para esta interacción.')

    try:
        file_path = interaction.audio_file.path
        if os.path.exists(file_path):
            mime_type, _ = mimetypes.guess_type(file_path)
            return FileResponse(
                open(file_path, 'rb'),
                content_type=mime_type or 'audio/mpeg',
            )
    except (NotImplementedError, ValueError, AttributeError):
        pass  # Cloudinary no soporta .path — continúa al proxy

    try:
        with urllib.request.urlopen(interaction.audio_file.url, timeout=30) as remote:
            content      = remote.read()
            content_type = remote.headers.get_content_type() or 'audio/mpeg'
        filename = os.path.basename(interaction.audio_file.name)
        response = HttpResponse(content, content_type=content_type)
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
    except Exception as exc:
        logger.error('Error sirviendo grabacion %s: %s', interaction_id, exc)
        raise Http404('No se pudo acceder al archivo de grabacion.')

# ─── HU-39: Métricas de tiempos de atención ──────────────────────────────────

def _build_attention_time_queryset(request):
    """
    Helper compartido por las tres vistas de HU-39.
    Retorna casos completos con deadline, aplicando filtros desde/hasta/sala.
    """
    desde_raw = (request.GET.get('desde') or '').strip()
    hasta_raw = (request.GET.get('hasta') or '').strip()
    sala_raw  = (request.GET.get('sala')  or '').strip()
    valid_salas = {value for value, _ in Case.ROOM_CHOICES}

    cases = (
        Case.objects
        .filter(
            status=Case.STATUS_COMPLETE,
            deadline_date__isnull=False,
        )
        .order_by('created_at')
    )

    if _parse_report_date(desde_raw):
        cases = cases.filter(created_at__date__gte=_parse_report_date(desde_raw))
    if _parse_report_date(hasta_raw):
        cases = cases.filter(created_at__date__lte=_parse_report_date(hasta_raw))
    if sala_raw in valid_salas:
        cases = cases.filter(sala=sala_raw)

    return cases, desde_raw, hasta_raw, sala_raw if sala_raw in valid_salas else ''


def _compute_attention_metrics(cases):
    """
    Calcula métricas de tiempo a partir de un queryset de casos con deadline.
    """
    today = timezone.localdate()
    rows  = []
    total_days = 0

    for case in cases:
        days_assigned = (case.deadline_date - case.created_at.date()).days
        days_remaining = (case.deadline_date - today).days
        overdue = days_remaining < 0

        rows.append({
            'code':           case.code,
            'sala':           case.get_sala_display() if case.sala else '—',
            'state':          case.state,
            'created_at':     case.created_at.strftime('%d/%m/%Y'),
            'deadline_date':  case.deadline_date.strftime('%d/%m/%Y'),
            'days_assigned':  days_assigned,
            'days_remaining': days_remaining,
            'overdue':        overdue,
        })
        total_days += days_assigned

    total     = len(rows)
    overdue   = sum(1 for r in rows if r['overdue'])
    on_time   = total - overdue
    avg_days  = round(total_days / total, 1) if total else 0.0

    return rows, total, overdue, on_time, avg_days


@role_required(ROLE_ADMINISTRADOR, ROLE_PROFESOR)
def case_attention_time_report(request):
    """HU-39: Métricas de tiempos de atención."""
    cases, desde_raw, hasta_raw, sala_filter = _build_attention_time_queryset(request)
    rows, total, overdue, on_time, avg_days  = _compute_attention_metrics(cases)

    return render(request, 'cases/attention_time_report.html', {
        'page_title':    'Métricas de tiempos de atención',
        'rows':          rows,
        'total':         total,
        'overdue':       overdue,
        'on_time':       on_time,
        'avg_days':      avg_days,
        'filtro_desde':  desde_raw,
        'filtro_hasta':  hasta_raw,
        'filtro_sala':   sala_filter,
        'salas':         Case.ROOM_CHOICES,
        'chart_labels':  ['A tiempo', 'Vencidos'],
        'chart_values':  [on_time, overdue],
    })


@role_required(ROLE_ADMINISTRADOR, ROLE_PROFESOR)
def export_attention_time_excel(request):
    """HU-39: Exporta métricas de tiempos de atención a Excel."""
    cases, desde_raw, hasta_raw, sala_filter = _build_attention_time_queryset(request)
    rows, total, overdue, on_time, avg_days  = _compute_attention_metrics(cases)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tiempos de atención'

    # ── Encabezado resumen ──────────────────────────────────────────────────
    resumen_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    resumen_font = Font(color='FFFFFF', bold=True)
    for col, texto in enumerate(['Total casos', 'A tiempo', 'Vencidos', 'Promedio días'], start=1):
        c = ws.cell(row=1, column=col, value=texto)
        c.fill = resumen_fill
        c.font = resumen_font
        c.alignment = Alignment(horizontal='center')
    for col, val in enumerate([total, on_time, overdue, avg_days], start=1):
        ws.cell(row=2, column=col, value=val).alignment = Alignment(horizontal='center')

    # ── Tabla detalle ───────────────────────────────────────────────────────
    headers = ['Código', 'Sala', 'Estado', 'Fecha creación',
               'Fecha límite', 'Días asignados', 'Días restantes', 'Vencido']
    header_fill = PatternFill(start_color='2E6DA4', end_color='2E6DA4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, start=5):
        ws.cell(row=row_idx, column=1).value = row['code']
        ws.cell(row=row_idx, column=2).value = row['sala']
        ws.cell(row=row_idx, column=3).value = row['state']
        ws.cell(row=row_idx, column=4).value = row['created_at']
        ws.cell(row=row_idx, column=5).value = row['deadline_date']
        ws.cell(row=row_idx, column=6).value = row['days_assigned']
        ws.cell(row=row_idx, column=7).value = row['days_remaining']
        ws.cell(row=row_idx, column=8).value = 'Sí' if row['overdue'] else 'No'
        if row['overdue']:
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).font = Font(color='CC0000')

    for col, width in enumerate([15, 12, 38, 15, 15, 16, 16, 10], start=1):
        ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_tiempos_atencion.xlsx"'
    return response


@role_required(ROLE_ADMINISTRADOR, ROLE_PROFESOR)
def export_attention_time_pdf(request):
    """HU-39: Exporta métricas de tiempos de atención a PDF."""
    cases, desde_raw, hasta_raw, sala_filter = _build_attention_time_queryset(request)
    rows, total, overdue, on_time, avg_days  = _compute_attention_metrics(cases)

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5 * inch, leftMargin=0.5 * inch,
        topMargin=0.5 * inch,   bottomMargin=0.5 * inch,
    )

    styles   = getSampleStyleSheet()
    elements = [
        Paragraph('<b>Métricas de Tiempos de Atención — Consultorio Jurídico ICESI</b>', styles['Title']),
        Spacer(1, 0.15 * inch),
        Paragraph(
            f'Total: {total} | A tiempo: {on_time} | Vencidos: {overdue} | '
            f'Promedio días asignados: {avg_days}',
            styles['Normal'],
        ),
        Spacer(1, 0.15 * inch),
    ]

    data = [['Código', 'Sala', 'Estado', 'F. Creación', 'F. Límite', 'Días asig.', 'Días rest.', 'Vencido']]
    for row in rows:
        data.append([
            row['code'],
            row['sala'],
            row['state'],
            row['created_at'],
            row['deadline_date'],
            str(row['days_assigned']),
            str(row['days_remaining']),
            'Sí' if row['overdue'] else 'No',
        ])

    table = Table(
        data,
        colWidths=[1.1*inch, 0.9*inch, 2.2*inch, 1.0*inch, 1.0*inch, 0.8*inch, 0.8*inch, 0.7*inch],
    )
    table.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#1A3A5C')),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0, 0), (-1, 0),  8),
        ('ALIGN',          (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',       (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F6F9')]),
        ('GRID',           (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING',     (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 3),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_tiempos_atencion.pdf"'
    return response

# ─── WebRTC nativo (reemplaza VideoSDK) ─────────────────────────────────────

def get_ice_servers(request, case_id):
    api_key = os.environ.get('METERED_API_KEY', '')
    domain  = os.environ.get('METERED_DOMAIN', '')
    if not api_key or not domain:
        return JsonResponse({'error': 'TURN no configurado'}, status=503)
    url = f'https://{domain}/api/v1/turn/credentials?apiKey={api_key}'
    try:
        with urllib.request.urlopen(url, timeout=5) as resp:
            ice_servers = json.loads(resp.read())
        return JsonResponse({'iceServers': ice_servers})
    except Exception as exc:
        logger.error('Error obteniendo ICE servers: %s', exc)
        return JsonResponse({'error': 'No se pudo obtener servidores TURN'}, status=503)


@login_required
def create_call_session(request, case_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    case = get_object_or_404(Case, pk=case_id)
    if not can_add_interaction(request.user, case):
        return JsonResponse({'error': 'Sin permiso'}, status=403)
    session = CallSession.objects.create(case=case, created_by=request.user)
    return JsonResponse({'roomId': str(session.room_id)})


@login_required
def set_call_offer(request, case_id, room_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    session = get_object_or_404(CallSession, room_id=room_id, case_id=case_id)
    data = json.loads(request.body)
    session.offer_sdp = json.dumps(data['sdp'])
    session.save(update_fields=['offer_sdp'])
    return JsonResponse({'status': 'ok'})


def get_call_state(request, case_id, room_id):
    session = get_object_or_404(CallSession, room_id=room_id, case_id=case_id)
    return JsonResponse({
        'status':  session.status,
        'hasOffer': bool(session.offer_sdp),
        'answer':  json.loads(session.answer_sdp) if session.answer_sdp else None,
    })


@csrf_exempt
def set_call_answer(request, case_id, room_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    session = get_object_or_404(CallSession, room_id=room_id, case_id=case_id)
    data = json.loads(request.body)
    session.answer_sdp = json.dumps(data['sdp'])
    session.status = CallSession.STATUS_ACTIVE
    session.save(update_fields=['answer_sdp', 'status'])
    return JsonResponse({'status': 'ok'})


def get_call_offer(request, case_id, room_id):
    session = get_object_or_404(CallSession, room_id=room_id, case_id=case_id)
    return JsonResponse({
        'offer':  json.loads(session.offer_sdp) if session.offer_sdp else None,
        'status': session.status,
    })


def join_webrtc_call(request, case_id, room_id):
    case = get_object_or_404(Case, pk=case_id)
    get_object_or_404(CallSession, room_id=room_id, case_id=case_id)
    return render(request, 'cases/call_room.html', {
        'case': case,
        'room_id': room_id,
        'ice_url': request.build_absolute_uri(f'/casos/{case_id}/webrtc/ice/'),
        'offer_url': request.build_absolute_uri(f'/casos/{case_id}/webrtc/{room_id}/oferta/leer/'),
        'answer_url': request.build_absolute_uri(f'/casos/{case_id}/webrtc/{room_id}/respuesta/'),
        'state_url': request.build_absolute_uri(f'/casos/{case_id}/webrtc/{room_id}/estado/'),
    })

