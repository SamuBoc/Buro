import io
from datetime import date, datetime

from django.contrib import messages
from django.db.models import Count
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from accounts.constants import ROLE_ADMINISTRADOR, ROLE_SECRETARIA
from accounts.decorators import role_required
from beneficiary.models import Beneficiary
from beneficiary.signals import log_beneficiary_cite_attendance
from core.utils import get_client_ip

from .forms import CiteForm, RescheduleCiteForm
from .models import Cite


@login_required
def create_cite(request, beneficiary_id):
    beneficiary = get_object_or_404(Beneficiary, pk=beneficiary_id)

    if request.method == 'POST':
        form = CiteForm(request.POST)
        if form.is_valid():
            cite            = form.save(commit=False)
            cite.beneficiary = beneficiary
            cite.save()
            messages.success(request, 'Cita agendada correctamente con la modalidad seleccionada.')
            return redirect('beneficiary_detail', pk=beneficiary.id)
        messages.error(request, 'Debes seleccionar una modalidad valida para continuar.')
    else:
        form = CiteForm()

    return render(request, 'cite/cite_form.html', {
        'form':        form,
        'beneficiary': beneficiary,
    })


@login_required
def beneficiary_cites(request, beneficiary_id):
    beneficiary = get_object_or_404(Beneficiary, pk=beneficiary_id)
    cites       = Cite.objects.filter(beneficiary=beneficiary).order_by('-id')

    return render(request, 'cite/cite_list.html', {
        'beneficiary': beneficiary,
        'cites':       cites,
    })


@login_required
def reschedule_cite(request, pk):
    cite = get_object_or_404(Cite, pk=pk)

    if request.method == 'POST':
        form = RescheduleCiteForm(request.POST, instance=cite)
        if form.is_valid():
            form.save()
            return redirect('beneficiary_cites', beneficiary_id=cite.beneficiary_id)
        messages.error(request, 'Corrige los errores del formulario')
    else:
        form = RescheduleCiteForm(instance=cite)

    return render(request, 'cite/reschedule_cite.html', {
        'form': form,
        'cite': cite,
    })


@login_required
def cancel_cite(request, pk):
    cite = get_object_or_404(Cite, pk=pk)
    if request.method == 'POST':
        cite.state_cite = Cite.STATE_CANCELED
        cite.save()
    return redirect('beneficiary_cites', beneficiary_id=cite.beneficiary_id)


@login_required
def register_cite_attendance(request, pk, status):
    cite = get_object_or_404(Cite, pk=pk)

    if request.method != 'POST':
        return redirect('beneficiary_cites', beneficiary_id=cite.beneficiary_id)

    if cite.state_cite == Cite.STATE_CANCELED:
        messages.error(request, 'No puedes registrar asistencia en una cita cancelada.')
        return redirect('beneficiary_cites', beneficiary_id=cite.beneficiary_id)

    if cite.date_assigned and cite.date_assigned > date.today():
        messages.warning(request, 'Solo puedes registrar asistencia en citas de hoy o anteriores.')
        return redirect('beneficiary_cites', beneficiary_id=cite.beneficiary_id)

    status_map = {
        'asistio':    (Cite.STATE_ATTENDED, True),
        'no-asistio': (Cite.STATE_NO_SHOW,  False),
    }

    if status not in status_map:
        messages.error(request, 'Estado de asistencia no válido.')
        return redirect('beneficiary_cites', beneficiary_id=cite.beneficiary_id)

    new_state, attended = status_map[status]

    if cite.state_cite != new_state:
        cite.state_cite = new_state
        cite.save()
        log_beneficiary_cite_attendance(
            cite.beneficiary,
            cite,
            request.user,
            attended,
            ip=get_client_ip(request),
        )
        messages.success(request, 'Asistencia registrada correctamente.')
    else:
        messages.info(request, 'La cita ya tiene este estado registrado.')

    return redirect('beneficiary_cites', beneficiary_id=cite.beneficiary_id)


@role_required(ROLE_ADMINISTRADOR, ROLE_SECRETARIA)
def cite_report(request):
    """HU-19: Reporte de citas no confirmadas o no asistidas."""
    desde_raw  = (request.GET.get('desde')  or '').strip()
    hasta_raw  = (request.GET.get('hasta')  or '').strip()
    estado_raw = (request.GET.get('estado') or '').strip()

    estados_validos = [Cite.STATE_PENDING, Cite.STATE_CANCELED]
    estado_filter   = estado_raw if estado_raw in estados_validos else ''

    cites = Cite.objects.select_related('beneficiary').filter(
        state_cite__in=estados_validos
    )

    if desde_raw:
        try:
            cites = cites.filter(
                date_assigned__gte=datetime.strptime(desde_raw, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    if hasta_raw:
        try:
            cites = cites.filter(
                date_assigned__lte=datetime.strptime(hasta_raw, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    if estado_filter:
        cites = cites.filter(state_cite=estado_filter)

    cites = cites.order_by('date_assigned')

    return render(request, 'cite/cite_report.html', {
        'cites':         cites,
        'total':         cites.count(),
        'filtro_desde':  desde_raw,
        'filtro_hasta':  hasta_raw,
        'filtro_estado': estado_filter,
        'estados':       estados_validos,
        'page_title':    'Reporte de citas no confirmadas o no asistidas',
    })


@role_required(ROLE_ADMINISTRADOR, ROLE_SECRETARIA)
def cite_report_excel(request):
    """HU-19: Exporta el reporte de citas a Excel."""
    cites = Cite.objects.select_related('beneficiary').filter(
        state_cite__in=[Cite.STATE_PENDING, Cite.STATE_CANCELED]
    ).order_by('date_assigned')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Citas'

    headers     = ['ID Cita', 'Beneficiario', 'Email', 'Fecha', 'Modalidad', 'Estado', 'Canal']
    header_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, cite in enumerate(cites, start=2):
        ws.cell(row=row_idx, column=1).value = cite.id
        ws.cell(row=row_idx, column=2).value = cite.beneficiary.name
        ws.cell(row=row_idx, column=3).value = cite.beneficiary.email
        ws.cell(row=row_idx, column=4).value = cite.date_assigned.strftime('%d/%m/%Y')
        ws.cell(row=row_idx, column=5).value = cite.get_modality_cite_display()
        ws.cell(row=row_idx, column=6).value = cite.state_cite
        ws.cell(row=row_idx, column=7).value = cite.request_cite

    col_widths = [10, 25, 28, 14, 14, 14, 18]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_citas.xlsx"'
    return response


@role_required(ROLE_ADMINISTRADOR, ROLE_SECRETARIA)
def cite_report_pdf(request):
    """HU-19: Exporta el reporte de citas a PDF."""
    cites = Cite.objects.select_related('beneficiary').filter(
        state_cite__in=[Cite.STATE_PENDING, Cite.STATE_CANCELED]
    ).order_by('date_assigned')

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5 * inch,
        leftMargin=0.5  * inch,
        topMargin=0.5   * inch,
        bottomMargin=0.5 * inch,
    )

    styles   = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(
        '<b>Reporte de Citas No Confirmadas o No Asistidas — Buró Jurídico ICESI</b>',
        styles['Title'],
    ))
    elements.append(Spacer(1, 0.2 * inch))

    data = [['ID', 'Beneficiario', 'Email', 'Fecha', 'Modalidad', 'Estado']]
    for cite in cites:
        data.append([
            str(cite.id),
            cite.beneficiary.name,
            cite.beneficiary.email,
            cite.date_assigned.strftime('%d/%m/%Y'),
            cite.get_modality_cite_display(),
            cite.state_cite,
        ])

    table = Table(
        data,
        colWidths=[0.6*inch, 2.2*inch, 2.5*inch, 1.1*inch, 1.2*inch, 1.2*inch],
    )
    table.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#1A3A5C')),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0, 0), (-1, 0),  9),
        ('ALIGN',          (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',       (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F6F9')]),
        ('GRID',           (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING',     (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_citas.pdf"'
    return response


@role_required(ROLE_ADMINISTRADOR, ROLE_SECRETARIA)
def cite_attendance_report(request):
    """HU-38: Metricas de asistencia de citas."""
    desde_raw = (request.GET.get('desde') or '').strip()
    hasta_raw = (request.GET.get('hasta') or '').strip()

    cites = Cite.objects.filter(state_cite__in=[Cite.STATE_ATTENDED, Cite.STATE_NO_SHOW])

    if desde_raw:
        try:
            cites = cites.filter(
                date_assigned__gte=datetime.strptime(desde_raw, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    if hasta_raw:
        try:
            cites = cites.filter(
                date_assigned__lte=datetime.strptime(hasta_raw, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    counts = {
        row['state_cite']: row['count']
        for row in cites.values('state_cite').annotate(count=Count('id'))
    }

    attended_count = counts.get(Cite.STATE_ATTENDED, 0)
    no_show_count = counts.get(Cite.STATE_NO_SHOW, 0)
    total = attended_count + no_show_count

    attendance_percentage = round((attended_count / total) * 100, 1) if total else 0.0
    no_show_percentage = round((no_show_count / total) * 100, 1) if total else 0.0

    rows = [
        {
            'estado': Cite.STATE_ATTENDED,
            'cantidad': attended_count,
            'porcentaje': attendance_percentage,
        },
        {
            'estado': Cite.STATE_NO_SHOW,
            'cantidad': no_show_count,
            'porcentaje': no_show_percentage,
        },
    ]

    chart_labels = [row['estado'] for row in rows]
    chart_values = [row['cantidad'] for row in rows]

    return render(request, 'cite/cite_attendance_report.html', {
        'page_title': 'Metricas de asistencia a citas',
        'rows': rows,
        'total': total,
        'attended_count': attended_count,
        'no_show_count': no_show_count,
        'attendance_percentage': attendance_percentage,
        'no_show_percentage': no_show_percentage,
        'filtro_desde': desde_raw,
        'filtro_hasta': hasta_raw,
        'chart_labels': chart_labels,
        'chart_values': chart_values,
    })


@role_required(ROLE_ADMINISTRADOR, ROLE_SECRETARIA)
def cite_attendance_report_excel(request):
    """HU-38: Exporta las metricas de asistencia a Excel."""
    desde_raw = (request.GET.get('desde') or '').strip()
    hasta_raw = (request.GET.get('hasta') or '').strip()

    cites = Cite.objects.filter(state_cite__in=[Cite.STATE_ATTENDED, Cite.STATE_NO_SHOW])

    if desde_raw:
        try:
            cites = cites.filter(
                date_assigned__gte=datetime.strptime(desde_raw, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    if hasta_raw:
        try:
            cites = cites.filter(
                date_assigned__lte=datetime.strptime(hasta_raw, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    counts = {
        row['state_cite']: row['count']
        for row in cites.values('state_cite').annotate(count=Count('id'))
    }

    attended_count = counts.get(Cite.STATE_ATTENDED, 0)
    no_show_count = counts.get(Cite.STATE_NO_SHOW, 0)
    total = attended_count + no_show_count

    attendance_percentage = round((attended_count / total) * 100, 1) if total else 0.0
    no_show_percentage = round((no_show_count / total) * 100, 1) if total else 0.0

    wb = Workbook()
    ws = wb.active
    ws.title = 'Metricas'

    headers = ['Estado', 'Cantidad', 'Porcentaje']
    header_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    rows = [
        (Cite.STATE_ATTENDED, attended_count, f'{attendance_percentage}%'),
        (Cite.STATE_NO_SHOW, no_show_count, f'{no_show_percentage}%'),
        ('Total', total, '100%' if total else '0%'),
    ]

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1).value = row[0]
        ws.cell(row=row_idx, column=2).value = row[1]
        ws.cell(row=row_idx, column=3).value = row[2]

    col_widths = [22, 12, 14]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_asistencia.xlsx"'
    return response


@role_required(ROLE_ADMINISTRADOR, ROLE_SECRETARIA)
def cite_attendance_report_pdf(request):
    """HU-38: Exporta las metricas de asistencia a PDF."""
    desde_raw = (request.GET.get('desde') or '').strip()
    hasta_raw = (request.GET.get('hasta') or '').strip()

    cites = Cite.objects.filter(state_cite__in=[Cite.STATE_ATTENDED, Cite.STATE_NO_SHOW])

    if desde_raw:
        try:
            cites = cites.filter(
                date_assigned__gte=datetime.strptime(desde_raw, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    if hasta_raw:
        try:
            cites = cites.filter(
                date_assigned__lte=datetime.strptime(hasta_raw, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    counts = {
        row['state_cite']: row['count']
        for row in cites.values('state_cite').annotate(count=Count('id'))
    }

    attended_count = counts.get(Cite.STATE_ATTENDED, 0)
    no_show_count = counts.get(Cite.STATE_NO_SHOW, 0)
    total = attended_count + no_show_count

    attendance_percentage = round((attended_count / total) * 100, 1) if total else 0.0
    no_show_percentage = round((no_show_count / total) * 100, 1) if total else 0.0

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(
        '<b>Reporte de Metricas de Asistencia — Buró Juridico ICESI</b>',
        styles['Title'],
    ))
    elements.append(Spacer(1, 0.2 * inch))

    data = [
        ['Estado', 'Cantidad', 'Porcentaje'],
        [Cite.STATE_ATTENDED, str(attended_count), f'{attendance_percentage}%'],
        [Cite.STATE_NO_SHOW, str(no_show_count), f'{no_show_percentage}%'],
        ['Total', str(total), '100%' if total else '0%'],
    ]

    table = Table(data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0), colors.HexColor('#1A3A5C')),
        ('TEXTCOLOR',      (0, 0), (-1, 0), colors.white),
        ('FONTNAME',       (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',       (0, 0), (-1, 0), 9),
        ('ALIGN',          (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',       (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F6F9')]),
        ('GRID',           (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING',     (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_asistencia.pdf"'
    return response