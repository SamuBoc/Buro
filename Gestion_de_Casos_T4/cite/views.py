from datetime import date

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render

from beneficiary.models import Beneficiary
from beneficiary.signals import log_beneficiary_cite_attendance

from .forms import CiteForm, Reschedule_Cite
from .models import Cite


@login_required
def create_cite(request, beneficiary_id):
	beneficiary = get_object_or_404(Beneficiary, pk=beneficiary_id)

	if request.method == 'POST':
		form = CiteForm(request.POST)
		if form.is_valid():
			cite = form.save(commit=False)
			cite.beneficiary = beneficiary
			cite.save()
			messages.success(request, 'Cita agendada correctamente con la modalidad seleccionada.')
			return redirect('beneficiary_detail', pk=beneficiary.id)
		messages.error(request, 'Debes seleccionar una modalidad valida para continuar.')
	else:
		form = CiteForm()

	return render(request, 'schedule/cite_form.html', {
		'form': form,
		'beneficiary': beneficiary,
	})


@login_required
def beneficiary_cites(request, beneficiary_id):
	beneficiary = get_object_or_404(Beneficiary, pk=beneficiary_id)
	cites = Cite.objects.filter(beneficiary=beneficiary).order_by('-id')

	return render(request, 'schedule/cite_list.html', {
		'beneficiary': beneficiary,
		'cites': cites,
	})

@login_required
def reschedule_cite(request, pk):
	cite = get_object_or_404(Cite, pk=pk)

	if request.method == 'POST':
		form = Reschedule_Cite(request.POST, instance=cite)

		if form.is_valid():
			form.save()
			return redirect('beneficiary_cites', beneficiary_id = cite.beneficiary_id)
		else:
			print(form.errors)
			messages.error(request, 'Corrige los errores del formulario')
	
	else: form = Reschedule_Cite(instance=cite)
	return render(request, 'reschedule/reschedule_cite.html', {
		'form': form,
		'cite': cite
	})

@login_required
def cancel_cite(request, pk):
	if request.method == 'POST':
		cite = get_object_or_404(Cite, pk=pk)
		cite.state_cite = Cite.STATE_CANCELED
		cite.save()
		return redirect('beneficiary_cites', beneficiary_id = cite.beneficiary_id)


@login_required
def register_cite_attendance(request, pk, status):
	cite = get_object_or_404(Cite, pk=pk)

	if request.method != 'POST':
		return redirect('beneficiary_cites', beneficiary_id=cite.beneficiary_id)

	if cite.state_cite == Cite.STATE_CANCELED:
		messages.error(request, 'No puedes registrar asistencia en una cita cancelada.')
		return redirect('beneficiary_cites', beneficiary_id=cite.beneficiary_id)

	if cite.date_assigned and cite.date_assigned > date.today():
		messages.warning(request, 'Solo puedes registrar asistencia en citas de hoy o anteriores.')
		return redirect('beneficiary_cites', beneficiary_id=cite.beneficiary_id)

	status_map = {
		'asistio': (Cite.STATE_ATTENDED, True),
		'no-asistio': (Cite.STATE_NO_SHOW, False),
	}

	if status not in status_map:
		messages.error(request, 'Estado de asistencia no válido.')
		return redirect('beneficiary_cites', beneficiary_id=cite.beneficiary_id)

	new_state, attended = status_map[status]

	if cite.state_cite != new_state:
		cite.state_cite = new_state
		cite.save()
		log_beneficiary_cite_attendance(
			cite.beneficiary,
			cite,
			request.user,
			attended,
			ip=_get_client_ip(request),
		)
		messages.success(request, 'Asistencia registrada correctamente.')
	else:
		messages.info(request, 'La cita ya tiene este estado registrado.')

	return redirect('beneficiary_cites', beneficiary_id=cite.beneficiary_id)


def _get_client_ip(request):
	x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
	if x_forwarded_for:
		return x_forwarded_for.split(',')[0].strip()
	return request.META.get('REMOTE_ADDR')

import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

from accounts.constants import ROLE_ADMINISTRADOR, ROLE_SECRETARIA
from accounts.decorators import role_required


@role_required(ROLE_ADMINISTRADOR, ROLE_SECRETARIA)
def cite_report(request):
    """HU-19: Reporte de citas no confirmadas o no asistidas."""
    from datetime import datetime

    desde_raw = (request.GET.get('desde') or '').strip()
    hasta_raw = (request.GET.get('hasta') or '').strip()
    estado_raw = (request.GET.get('estado') or '').strip()

    estados_validos = [Cite.STATE_PENDING, Cite.STATE_CANCELED]
    estado_filter = estado_raw if estado_raw in estados_validos else ''

    cites = Cite.objects.select_related('beneficiary').filter(
        state_cite__in=estados_validos
    )

    if desde_raw:
        try:
            desde_date = datetime.strptime(desde_raw, '%Y-%m-%d').date()
            cites = cites.filter(date_assigned__gte=desde_date)
        except ValueError:
            pass

    if hasta_raw:
        try:
            hasta_date = datetime.strptime(hasta_raw, '%Y-%m-%d').date()
            cites = cites.filter(date_assigned__lte=hasta_date)
        except ValueError:
            pass

    if estado_filter:
        cites = cites.filter(state_cite=estado_filter)

    cites = cites.order_by('date_assigned')

    return render(request, 'schedule/cite_report.html', {
        'cites':         cites,
        'total':         cites.count(),
        'filtro_desde':  desde_raw,
        'filtro_hasta':  hasta_raw,
        'filtro_estado': estado_filter,
        'estados':       estados_validos,
        'page_title':    'Reporte de citas no confirmadas o no asistidas',
    })


@role_required(ROLE_ADMINISTRADOR, ROLE_SECRETARIA)
def cite_report_excel(request):
    """HU-19: Exporta el reporte de citas a Excel."""
    cites = Cite.objects.select_related('beneficiary').filter(
        state_cite__in=[Cite.STATE_PENDING, Cite.STATE_CANCELED]
    ).order_by('date_assigned')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Citas'

    headers = ['ID Cita', 'Beneficiario', 'Email', 'Fecha', 'Modalidad', 'Estado', 'Canal']
    header_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, cite in enumerate(cites, start=2):
        ws.cell(row=row_idx, column=1).value = cite.id
        ws.cell(row=row_idx, column=2).value = cite.beneficiary.name
        ws.cell(row=row_idx, column=3).value = cite.beneficiary.email
        ws.cell(row=row_idx, column=4).value = cite.date_assigned.strftime('%d/%m/%Y')
        ws.cell(row=row_idx, column=5).value = cite.get_modality_cite_display()
        ws.cell(row=row_idx, column=6).value = cite.state_cite
        ws.cell(row=row_idx, column=7).value = cite.request_cite

    col_widths = [10, 25, 28, 14, 14, 14, 18]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_citas.xlsx"'
    return response


@role_required(ROLE_ADMINISTRADOR, ROLE_SECRETARIA)
def cite_report_pdf(request):
    """HU-19: Exporta el reporte de citas a PDF."""
    cites = Cite.objects.select_related('beneficiary').filter(
        state_cite__in=[Cite.STATE_PENDING, Cite.STATE_CANCELED]
    ).order_by('date_assigned')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph(
        '<b>Reporte de Citas No Confirmadas o No Asistidas — Buró Jurídico ICESI</b>',
        styles['Title']
    )
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))

    data = [['ID', 'Beneficiario', 'Email', 'Fecha', 'Modalidad', 'Estado']]

    for cite in cites:
        data.append([
            str(cite.id),
            cite.beneficiary.name,
            cite.beneficiary.email,
            cite.date_assigned.strftime('%d/%m/%Y'),
            cite.get_modality_cite_display(),
            cite.state_cite,
        ])

    table = Table(
        data,
        colWidths=[0.6*inch, 2.2*inch, 2.5*inch, 1.1*inch, 1.2*inch, 1.2*inch]
    )
    table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1A3A5C')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  9),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F6F9')]),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_citas.pdf"'
    return response