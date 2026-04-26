from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from accounts.constants import ROLE_ADMINISTRADOR, ROLE_PROFESOR, ROLE_SECRETARIA
from accounts.decorators import role_required
from accounts.permissions import can_manage_case_deadline, can_reassign_case, can_view_case

from .forms import CaseDeadlineForm, CaseForm, CaseReassignmentForm, CaseRejectionForm
from .models import Case, CaseAuditLog, CaseDocument, Notification
from .services import auto_assign_case, reassign_case


def _get_user_draft(user):
    return (
        Case.objects
        .filter(created_by=user, status=Case.STATUS_DRAFT)
        .order_by('-created_at', '-pk')
        .first()
    )


def _can_access_draft(user, case):
    return user.is_superuser or case.created_by_id == user.id


def _render_case_form(request, form, draft_case=None, is_editing_draft=False):
    return render(request, 'cases/case_register.html', {
        'form': form,
        'draft_case': draft_case,
        'is_editing_draft': is_editing_draft,
    })


def _build_deadline_priority(case, today):
    if not case.deadline_date:
        return {
            'text': 'Sin fecha limite',
            'days_remaining': None,
            'priority_label': 'Sin prioridad',
            'priority_class': 'priority-none',
            'deadline_class': 'deadline-none',
        }

    days_remaining = (case.deadline_date - today).days

    if days_remaining < 0:
        priority_label = 'Critica'
        priority_class = 'priority-critical'
        deadline_text = f'Vencido hace {abs(days_remaining)} dia(s)'
        deadline_class = 'deadline-overdue'
    elif days_remaining <= 2:
        priority_label = 'Alta'
        priority_class = 'priority-high'
        deadline_text = 'Vence hoy' if days_remaining == 0 else f'Vence en {days_remaining} dia(s)'
        deadline_class = 'deadline-soon'
    elif days_remaining <= 7:
        priority_label = 'Media'
        priority_class = 'priority-medium'
        deadline_text = f'Vence en {days_remaining} dia(s)'
        deadline_class = 'deadline-normal'
    else:
        priority_label = 'Baja'
        priority_class = 'priority-low'
        deadline_text = f'Vence en {days_remaining} dia(s)'
        deadline_class = 'deadline-normal'

    return {
        'text': deadline_text,
        'days_remaining': days_remaining,
        'priority_label': priority_label,
        'priority_class': priority_class,
        'deadline_class': deadline_class,
    }


@login_required
def case_list(request):
    today = timezone.localdate()
    cases = list(Case.objects.select_related('beneficiary', 'assigned_student').all())

    for case in cases:
        deadline_priority = _build_deadline_priority(case, today)
        case.deadline_status_text = deadline_priority['text']
        case.days_remaining = deadline_priority['days_remaining']
        case.priority_label = deadline_priority['priority_label']
        case.priority_class = deadline_priority['priority_class']
        case.deadline_class = deadline_priority['deadline_class']

    return render(request, 'cases/case_list.html', {
        'cases': cases,
    })


@role_required(ROLE_SECRETARIA, ROLE_ADMINISTRADOR)
def case_create(request):
    draft_case = _get_user_draft(request.user)

    if request.method == 'POST':
        submit_action = request.POST.get('submit_action', 'complete')
        is_draft_submission = submit_action == 'draft'
        form = CaseForm(
            request.POST,
            request.FILES,
            instance=draft_case,
            allow_partial=is_draft_submission,
        )

        if form.is_valid():
            try:
                with transaction.atomic():
                    case = form.save(commit=False)
                    case.created_by = request.user
                    case._request = request
                    case.status = (
                        Case.STATUS_DRAFT
                        if is_draft_submission
                        else Case.STATUS_COMPLETE
                    )
                    case.save()

                    for document in form.cleaned_data.get('documents', []):
                        CaseDocument.objects.create(case=case, file=document)

                    student = None
                    if case.status == Case.STATUS_COMPLETE:
                        student = auto_assign_case(case)

                if is_draft_submission:
                    messages.success(
                        request,
                        f'Se guardo el borrador del caso {case.code}. Puedes continuarlo despues.'
                    )
                elif student is None:
                    messages.warning(
                        request,
                        f'El caso {case.code} fue registrado pero no hay estudiantes disponibles para asignarlo.'
                    )
                else:
                    messages.success(
                        request,
                        f'El caso {case.code} fue registrado y asignado a {student.get_full_name() or student.username}.'
                    )
                if is_draft_submission:
                    return redirect('case_create')
                return redirect('case_detail', pk=case.pk)
            except Exception:
                messages.error(
                    request,
                    'Ocurrio un problema al guardar el caso y sus documentos. Intente nuevamente.'
                )
        else:
            messages.error(request, 'Por favor corrija los errores del formulario.')
    else:
        form = CaseForm(instance=draft_case, allow_partial=bool(draft_case))
        if draft_case:
            messages.info(
                request,
                f'Tienes un borrador pendiente ({draft_case.code}). Puedes continuar completandolo.'
            )

    return _render_case_form(request, form, draft_case=draft_case)


@login_required
def case_draft_list(request):
    drafts = (
        Case.objects
        .filter(created_by=request.user, status=Case.STATUS_DRAFT)
        .select_related('beneficiary', 'assigned_student')
        .order_by('-created_at')
    )

    return render(request, 'cases/case_draft_list.html', {
        'drafts': drafts,
    })


@login_required
def case_edit_draft(request, pk):
    draft_case = get_object_or_404(
        Case.objects.select_related('beneficiary', 'assigned_student'),
        pk=pk,
        status=Case.STATUS_DRAFT,
    )

    if not _can_access_draft(request.user, draft_case):
        messages.error(request, 'No tienes permisos para editar este borrador.')
        return redirect('case_draft_list')

    if request.method == 'POST':
        submit_action = request.POST.get('submit_action', 'draft')
        is_draft_submission = submit_action == 'draft'
        form = CaseForm(
            request.POST,
            request.FILES,
            instance=draft_case,
            allow_partial=is_draft_submission,
        )

        if form.is_valid():
            try:
                with transaction.atomic():
                    case = form.save(commit=False)
                    case.created_by = draft_case.created_by or request.user
                    case._request = request
                    case.status = (
                        Case.STATUS_DRAFT if is_draft_submission else Case.STATUS_COMPLETE
                    )
                    case.save()

                    for document in form.cleaned_data.get('documents', []):
                        CaseDocument.objects.create(case=case, file=document)

                    student = None
                    if case.status == Case.STATUS_COMPLETE:
                        student = auto_assign_case(case)

                if is_draft_submission:
                    messages.success(
                        request,
                        f'Se actualizo el borrador del caso {case.code}.'
                    )
                    return redirect('case_edit_draft', pk=case.pk)

                if student is None:
                    messages.warning(
                        request,
                        f'El caso {case.code} fue completado pero no hay estudiantes disponibles para asignarlo.'
                    )
                else:
                    messages.success(
                        request,
                        f'El caso {case.code} fue completado y asignado a {student.get_full_name() or student.username}.'
                    )
                return redirect('case_detail', pk=case.pk)
            except Exception:
                messages.error(
                    request,
                    'Ocurrio un problema al actualizar el borrador. Intente nuevamente.'
                )
        else:
            messages.error(request, 'Por favor corrija los errores del formulario.')
    else:
        form = CaseForm(instance=draft_case, allow_partial=True)

    return _render_case_form(
        request,
        form,
        draft_case=draft_case,
        is_editing_draft=True,
    )


@login_required
def case_detail(request, pk):
    case = get_object_or_404(
        Case.objects
        .select_related('beneficiary', 'assigned_student')
        .prefetch_related('documents', 'reassignment_logs__changed_by', 'reassignment_logs__old_student', 'reassignment_logs__new_student'),
        pk=pk
    )

    if not can_view_case(request.user, case):
        messages.error(request, 'No tienes permisos para acceder a este caso.')
        return redirect('case_list')

    return render(request, 'cases/case_detail.html', {
        'case': case,
        'can_reassign': can_reassign_case(request.user),
        'can_manage_deadline': can_manage_case_deadline(request.user),
        'deadline_form': CaseDeadlineForm(instance=case),
        'reassignment_form': CaseReassignmentForm(case=case),
        'rejection_form': CaseRejectionForm(instance=case),
    })


@role_required(ROLE_SECRETARIA, ROLE_PROFESOR, ROLE_ADMINISTRADOR)
def case_update_deadline(request, pk):
    case = get_object_or_404(
        Case.objects.select_related('beneficiary', 'assigned_student'),
        pk=pk
    )

    if request.method != 'POST':
        return redirect('case_detail', pk=case.pk)

    form = CaseDeadlineForm(request.POST, instance=case)

    if form.is_valid():
        deadline_case = form.save()
        messages.success(
            request,
            f'La fecha limite del caso {deadline_case.code} fue actualizada.'
        )
        return redirect('case_detail', pk=case.pk)

    return render(request, 'cases/case_detail.html', {
        'case': case,
        'can_reassign': can_reassign_case(request.user),
        'can_manage_deadline': can_manage_case_deadline(request.user),
        'deadline_form': form,
        'reassignment_form': CaseReassignmentForm(case=case),
        'rejection_form': CaseRejectionForm(instance=case),
    })


@role_required(ROLE_SECRETARIA, ROLE_PROFESOR, ROLE_ADMINISTRADOR)
def case_reassign(request, pk):
    case = get_object_or_404(
        Case.objects.select_related('beneficiary', 'assigned_student'),
        pk=pk
    )

    if request.method != 'POST':
        return redirect('case_detail', pk=case.pk)

    form = CaseReassignmentForm(request.POST, case=case)

    if form.is_valid():
        new_student = form.cleaned_data['assigned_student']
        old_student = reassign_case(case, new_student, request.user)
        previous_name = old_student.get_full_name() or old_student.username if old_student else 'Sin asignar'
        new_name = new_student.get_full_name() or new_student.username
        messages.success(
            request,
            f'El caso {case.code} fue reasignado de {previous_name} a {new_name}.'
        )
        return redirect('case_detail', pk=case.pk)

    messages.error(request, 'Por favor seleccione un estudiante valido.')
    return render(request, 'cases/case_detail.html', {
        'case': case,
        'can_reassign': can_reassign_case(request.user),
        'can_manage_deadline': can_manage_case_deadline(request.user),
        'deadline_form': CaseDeadlineForm(instance=case),
        'reassignment_form': form,
        'rejection_form': CaseRejectionForm(instance=case),
    })


@role_required(ROLE_SECRETARIA, ROLE_PROFESOR, ROLE_ADMINISTRADOR)
def case_reject(request, pk):
    case = get_object_or_404(
        Case.objects.select_related('beneficiary', 'assigned_student'),
        pk=pk
    )

    if request.method != 'POST':
        return redirect('case_detail', pk=case.pk)

    form = CaseRejectionForm(request.POST, instance=case)

    if form.is_valid():
        try:
            with transaction.atomic():
                case.state = Case.STATE_REJECTED
                case.rejection_reason = form.cleaned_data['rejection_reason']
                case._request = request
                case.save()

                messages.success(
                    request,
                    f'El caso {case.code} ha sido rechazado exitosamente.'
                )
        except Exception:
            messages.error(
                request,
                'Ocurrió un problema al rechazar el caso. Intente nuevamente.'
            )
        return redirect('case_detail', pk=case.pk)

    messages.error(request, 'Por favor ingrese una causal de rechazo válida.')
    return render(request, 'cases/case_detail.html', {
        'case': case,
        'can_reassign': can_reassign_case(request.user),
        'can_manage_deadline': can_manage_case_deadline(request.user),
        'deadline_form': CaseDeadlineForm(instance=case),
        'reassignment_form': CaseReassignmentForm(case=case),
        'rejection_form': form,
    })


@login_required
def case_audit_log(request, case_id):
    case = get_object_or_404(Case, pk=case_id)

    has_access = (
        request.user.is_staff
        or request.user.groups.filter(
            name__in=[ROLE_SECRETARIA, ROLE_PROFESOR, ROLE_ADMINISTRADOR]
        ).exists()
        or case.assigned_student == request.user
    )
    if not has_access:
        messages.error(request, 'No tienes permiso para ver la bitacora de este caso.')
        return redirect('case_list')

    logs = CaseAuditLog.objects.filter(case=case).select_related('user').order_by('-timestamp')

    return render(request, 'cases/case_audit_log.html', {
        'case': case,
        'logs': logs,
        'page_title': f'Bitacora - {case.code}',
    })


@login_required
def global_audit_log(request):
    if not (
        request.user.is_staff
        or request.user.groups.filter(name=ROLE_ADMINISTRADOR).exists()
    ):
        messages.error(request, 'Acceso restringido a administradores.')
        return redirect('case_list')

    logs = CaseAuditLog.objects.select_related('user', 'case').order_by('-timestamp')[:500]
    return render(request, 'cases/global_audit_log.html', {
        'logs': logs,
        'page_title': 'Bitacora Global de Casos',
    })


@login_required
def notification_list(request):
    notifications = Notification.objects.filter(
        recipient_user=request.user
    ).select_related('case').order_by('-created_at')

    unread_count = notifications.filter(is_read=False).count()

    return render(request, 'cases/notifications.html', {
        'notifications': notifications,
        'unread_count': unread_count,
        'page_title': 'Mis Notificaciones',
    })


@login_required
def mark_notification_read(request, notification_id):
    notification = get_object_or_404(
        Notification,
        pk=notification_id,
        recipient_user=request.user,
    )
    notification.mark_as_read()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'ok', 'notification_id': notification_id})

    return redirect('case_detail', pk=notification.case_id)


@login_required
def mark_all_notifications_read(request):
    if request.method == 'POST':
        Notification.objects.filter(
            recipient_user=request.user,
            is_read=False,
        ).update(is_read=True, read_at=timezone.now())
        messages.success(request, 'Todas las notificaciones marcadas como leidas.')
    return redirect('notification_list')


@login_required
def unread_notifications_count(request):
    count = Notification.objects.filter(
        recipient_user=request.user,
        is_read=False,
    ).count()
    return JsonResponse({'unread_count': count})


REPORT_KNOWN_STATES = [
    Case.STATE_PENDING,
    Case.STATE_ASSIGNED,
    Case.STATE_NO_STUDENTS,
    Case.STATE_REJECTED,
]


def _parse_report_date(raw_value):
    if not raw_value:
        return None
    try:
        return datetime.strptime(raw_value, '%Y-%m-%d').date()
    except ValueError:
        return None


@role_required(ROLE_ADMINISTRADOR, ROLE_PROFESOR)
def case_report_by_state(request):
    desde_raw = (request.GET.get('desde') or '').strip()
    hasta_raw = (request.GET.get('hasta') or '').strip()
    sala_raw = (request.GET.get('sala') or '').strip()

    desde_date = _parse_report_date(desde_raw)
    hasta_date = _parse_report_date(hasta_raw)

    valid_salas = {value for value, _ in Case.ROOM_CHOICES}
    sala_filter = sala_raw if sala_raw in valid_salas else ''

    cases = Case.objects.all()
    if desde_date:
        cases = cases.filter(created_at__date__gte=desde_date)
    if hasta_date:
        cases = cases.filter(created_at__date__lte=hasta_date)
    if sala_filter:
        cases = cases.filter(sala=sala_filter)

    total = cases.count()

    state_counts = {
        row['state']: row['count']
        for row in cases.values('state').annotate(count=Count('id'))
    }

    rows = []
    for state in REPORT_KNOWN_STATES:
        cantidad = state_counts.pop(state, 0)
        porcentaje = round((cantidad / total * 100), 1) if total else 0.0
        rows.append({
            'estado': state,
            'cantidad': cantidad,
            'porcentaje': porcentaje,
        })
    for extra_state, cantidad in state_counts.items():
        porcentaje = round((cantidad / total * 100), 1) if total else 0.0
        rows.append({
            'estado': extra_state or 'Sin estado',
            'cantidad': cantidad,
            'porcentaje': porcentaje,
        })

    chart_labels = [row['estado'] for row in rows]
    chart_values = [row['cantidad'] for row in rows]

    return render(request, 'cases/report_by_state.html', {
        'page_title': 'Reporte de casos por estado',
        'rows': rows,
        'total': total,
        'filtro_desde': desde_raw,
        'filtro_hasta': hasta_raw,
        'filtro_sala': sala_filter,
        'salas': Case.ROOM_CHOICES,
        'chart_labels': chart_labels,
        'chart_values': chart_values,
    })

import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch


@role_required(ROLE_ADMINISTRADOR)
def export_cases_excel(request):
    """HU-40: Exporta todos los casos a un archivo Excel."""
    cases = Case.objects.select_related(
        'beneficiary', 'assigned_student'
    ).order_by('-created_at')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Casos'

    headers = [
        'Código', 'Sala', 'Beneficiario', 'Estudiante Asignado',
        'Estado', 'Fecha de Creación', 'Fecha Límite'
    ]
    header_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, case in enumerate(cases, start=2):
        ws.cell(row=row_idx, column=1).value = case.code
        ws.cell(row=row_idx, column=2).value = case.get_sala_display()
        ws.cell(row=row_idx, column=3).value = case.beneficiary.name
        ws.cell(row=row_idx, column=4).value = (
            case.assigned_student.get_full_name() or case.assigned_student.username
            if case.assigned_student else 'Sin asignar'
        )
        ws.cell(row=row_idx, column=5).value = case.state
        ws.cell(row=row_idx, column=6).value = case.created_at.strftime('%d/%m/%Y')
        ws.cell(row=row_idx, column=7).value = (
            case.deadline_date.strftime('%d/%m/%Y') if case.deadline_date else '—'
        )

    column_widths = [15, 12, 25, 25, 35, 18, 15]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_casos.xlsx"'
    return response


@role_required(ROLE_ADMINISTRADOR)
def export_cases_pdf(request):
    """HU-40: Exporta todos los casos a un archivo PDF."""
    cases = Case.objects.select_related(
        'beneficiary', 'assigned_student'
    ).order_by('-created_at')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph(
        '<b>Reporte de Casos — Consultorio Jurídico ICESI</b>',
        styles['Title']
    )
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))

    data = [['Código', 'Sala', 'Beneficiario', 'Estudiante Asignado', 'Estado', 'Fecha Creación']]

    for case in cases:
        data.append([
            case.code,
            case.get_sala_display(),
            case.beneficiary.name,
            (
                case.assigned_student.get_full_name() or case.assigned_student.username
                if case.assigned_student else 'Sin asignar'
            ),
            case.state,
            case.created_at.strftime('%d/%m/%Y'),
        ])

    table = Table(data, colWidths=[1.2*inch, 1*inch, 2*inch, 2*inch, 2.5*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0),  colors.HexColor('#1A3A5C')),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0),  9),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',    (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F6F9')]),
        ('GRID',        (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING',  (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_casos.pdf"'
    return response